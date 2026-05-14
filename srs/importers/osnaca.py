"""
OSNACA workbook → SQL conversion pipeline.

Entry point: run_import(import_id)

Order of operations (matches the plan):
  1. Mark the ReferenceImport row as running
  2. Open both workbooks from the stored FileFields
  3. Seed Element / Mineral / DepositClassification if missing
  4. Upsert ReferenceDeposit rows from Deposit Locations
  5. Upsert ReferenceSample rows from Ore samples + CSIRO Cloncurry Samples
  6. Bulk-insert ReferenceSampleMeasurement rows from Data + Cloncurry Supplement + PGEs
  7. Record stats / errors, mark completed
"""
from __future__ import annotations

import re

import openpyxl
from django.db import transaction
from django.utils import timezone


_NUMERIC_RE = re.compile(r"-?\d+(?:\.\d+)?")


def _to_float(value) -> float | None:
    """
    Coerce a spreadsheet cell to a float, tolerating dirty data.
    Strips trailing junk like degree signs, asterisks, or stray whitespace.
    Returns None for blanks or unparseable values.
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = _NUMERIC_RE.search(str(value))
    return float(match.group(0)) if match else None

from ..models import (
    DepositClassification,
    Element,
    Mineral,
    ReferenceDeposit,
    ReferenceImport,
    ReferenceSample,
    ReferenceSampleMeasurement,
)
from .headers import flatten_headers, header_label
from .seed_data import ELEMENTS, ELEMENT_METHOD_OVERRIDES, NON_ELEMENT_COLUMNS


# top level entry

def run_import(import_id: int) -> ReferenceImport:
    # Ensure this thread has a fresh DB connection (required when called from a background thread)
    from django.db import close_old_connections
    close_old_connections()

    import_row = ReferenceImport.objects.get(pk=import_id)
    import_row.status = ReferenceImport.STATUS_RUNNING
    import_row.errors = []
    import_row.save(update_fields=["status", "errors"])

    stats = {"deposits": 0, "samples": 0, "measurements": 0, "warnings": 0}
    errors: list[dict] = []

    try:
        with transaction.atomic():
            data_wb = openpyxl.load_workbook(
                import_row.data_file.path, read_only=True, data_only=True
            )
            meta_wb = openpyxl.load_workbook(
                import_row.metadata_file.path, read_only=True, data_only=True
            )

            seed_lookups(meta_wb)
            deposits_by_code = load_deposits(meta_wb, import_row, stats)
            samples_by_code = load_samples(meta_wb, import_row, deposits_by_code, stats, errors)
            load_measurements(data_wb, import_row, samples_by_code, stats, errors)

        import_row.status = ReferenceImport.STATUS_COMPLETED
    except Exception as exc:
        import_row.status = ReferenceImport.STATUS_FAILED
        errors.append({"phase": "fatal", "reason": str(exc)})

    stats["warnings"] = len(errors)
    import_row.stats = stats
    import_row.errors = errors
    import_row.completed_at = timezone.now()
    import_row.save(update_fields=["status", "stats", "errors", "completed_at"])
    return import_row


# seed lookups

def seed_lookups(meta_wb) -> None:
    """Idempotent. Run on every import; no-op if rows already exist."""
    Element.objects.bulk_create(
        [
            Element(symbol=symbol, name=name, atomic_number=atomic_number, default_unit=default_unit)
            for symbol, name, atomic_number, default_unit in ELEMENTS
        ],
        update_conflicts=True,
        unique_fields=["symbol"],
        update_fields=["name", "atomic_number", "default_unit"],
    )

    ws = meta_wb["Mineral Codes"]
    Mineral.objects.bulk_create(
        [
            Mineral(code=code, name=name)
            for name, code, *_ in ws.iter_rows(min_row=2, values_only=True)
            if code and name
        ],
        update_conflicts=True,
        unique_fields=["code"],
        update_fields=["name"],
    )

    ws = meta_wb["Ore Deposit Classification"]
    current_class = None
    classification_objects = []
    for cls, sub, notes in ws.iter_rows(min_row=2, values_only=True):
        if cls:
            current_class = cls
        if current_class is None:
            continue
        classification_objects.append(DepositClassification(
            deposit_class=current_class,
            sub_class=sub or "",
            notes=notes or "",
        ))
    DepositClassification.objects.bulk_create(
        classification_objects,
        update_conflicts=True,
        unique_fields=["deposit_class", "sub_class"],
        update_fields=["notes"],
    )


# deposits

def load_deposits(meta_wb, import_row, stats) -> dict[str, ReferenceDeposit]:
    ws = meta_wb["Deposit Locations"]

    deposit_objects = []
    seen = set()
    for name, three_char, cls, sub, lng, lat in ws.iter_rows(min_row=2, values_only=True):
        if not name or not three_char or three_char in seen:
            continue
        seen.add(three_char)
        deposit_objects.append(ReferenceDeposit(
            name=name,
            three_char_code=three_char,
            import_ref=import_row,
            deposit_type=cls or "",
            mineral_system=sub or "",
            latitude=_to_float(lat),
            longitude=_to_float(lng),
            source="OSNACA",
        ))

    ReferenceDeposit.objects.bulk_create(
        deposit_objects,
        update_conflicts=True,
        unique_fields=["name", "three_char_code"],
        update_fields=["import_ref", "deposit_type", "mineral_system", "latitude", "longitude", "source"],
        batch_size=500,
    )

    three_char_codes = {obj.three_char_code for obj in deposit_objects}
    deposits_by_code = {
        d.three_char_code: d
        for d in ReferenceDeposit.objects.filter(three_char_code__in=three_char_codes)
    }

    stats["deposits"] = len(deposits_by_code)
    return deposits_by_code


# samples

def load_samples(meta_wb, import_row, deposits_by_code, stats, errors):
    # Pre-fetch all minerals once to avoid per-row DB queries
    minerals_by_code = {m.code: m.name for m in Mineral.objects.all()}

    sample_objects = []
    seen = set()
    for sheet_name in ("Ore samples", "CSIRO Cloncurry Samples"):
        ws = meta_wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            sample_code = row[0]
            three_char = row[4]
            if sample_code in (None, "") or str(sample_code) in seen:
                continue
            seen.add(str(sample_code))

            deposit = deposits_by_code.get(three_char)
            if deposit is None and three_char:
                errors.append({
                    "sheet": sheet_name, "row": row_idx,
                    "reason": f"Unknown three_char_code: {three_char!r}",
                })

            sample_objects.append(ReferenceSample(
                import_ref=import_row,
                sample_code=str(sample_code),
                reference_deposit=deposit,
                sample_type=f"{row[7] or ''} / {row[8] or ''}".strip(" /"),
                latitude=_to_float(row[11]),
                longitude=_to_float(row[10]),
                source_dataset="OSNACA",
                source_reference=row[57] or "",
                metadata=_build_sample_metadata(row, minerals_by_code),
            ))

    ReferenceSample.objects.bulk_create(
        sample_objects,
        update_conflicts=True,
        unique_fields=["import_ref", "sample_code"],
        update_fields=[
            "reference_deposit", "sample_type", "latitude", "longitude",
            "source_dataset", "source_reference", "metadata",
        ],
        batch_size=500,
    )

    samples_by_code = {
        s.sample_code: s
        for s in ReferenceSample.objects.filter(import_ref=import_row)
    }

    stats["samples"] = len(samples_by_code)
    return samples_by_code


def _build_sample_metadata(row, minerals_by_code: dict) -> dict:
    """Pack the long-tail columns into the JSON blob."""
    mineral_codes = [c for c in row[17:25] if c]
    minerals = [minerals_by_code[code] for code in mineral_codes if code in minerals_by_code]

    texture_headers = [
        "Massive-Homogenous", "Massive-Banded", "Massive-Granular",
        "Matrix or semi-massive", "Disseminated", "Blebs-Patches",
        "Stringer-Sulphide Veinlets", "Open space sulphide", "Breccia",
        "Sheared", "Vein Stockwork", "Vein(s) Crack-Seal", "Vein(s) Cockscomb",
        "Vein(s) Massive", "Vein(s) Shear Banded", "Vein(s) Breccia",
    ]
    textures = [
        label for label, cell in zip(texture_headers, row[25:41]) if cell == 1
    ]

    resource_headers = [
        "Mt", "Au_g_t", "Ag_g_t", "Pt_g_t", "Pd_g_t",
        "Cu_pct", "Zn_pct", "Pb_pct", "Ni_pct", "Co_pct",
        "Mo_pct", "W_pct", "Sn_pct", "U_pct", "Fe_pct", "Mn_pct",
    ]
    global_resource = {k: v for k, v in zip(resource_headers, row[41:57])}

    return {
        "donor": row[3],
        "your_sample_id": row[9],
        "utm_e": row[12], "utm_n": row[13],
        "rl": row[14], "srtm_rl": row[15],
        "description": row[16],
        "ore_minerals": minerals,
        "ore_textures": textures,
        "global_resource": global_resource,
    }


# measurements

def load_measurements(data_wb, import_row, samples_by_code, stats, errors):
    elements_by_symbol = {e.symbol: e for e in Element.objects.all()}
    rows_to_create: list[ReferenceSampleMeasurement] = []

    for sheet_name in ("Data", "Cloncurry Supplement", "PGEs"):
        if sheet_name not in data_wb.sheetnames:
            continue
        ws = data_wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 3:
            continue

        if sheet_name == "PGEs":
            headers = [("PGE", h or "") for h in all_rows[0]]
            data_start = 1
        else:
            headers = flatten_headers(all_rows[0], all_rows[1])
            data_start = 2

        for row in all_rows[data_start:]:
            sample_code = row[0]
            if sample_code in (None, ""):
                continue
            sample = samples_by_code.get(str(sample_code))
            if sample is None:
                errors.append({
                    "sheet": sheet_name,
                    "sample_code": sample_code,
                    "reason": "No matching ReferenceSample (missing in metadata workbook)",
                })
                continue

            for col_idx, (group, detail) in enumerate(headers):
                label = header_label(group, detail)
                if label in NON_ELEMENT_COLUMNS or label == "":
                    continue

                symbol, method = ELEMENT_METHOD_OVERRIDES.get(
                    (group, detail), (label, "")
                )
                element = elements_by_symbol.get(symbol)
                if element is None:
                    continue

                cell = row[col_idx]
                if cell in (None, ""):
                    continue
                if not isinstance(cell, (int, float)):
                    errors.append({
                        "sheet": sheet_name,
                        "sample_code": sample_code,
                        "column": label,
                        "reason": f"Non-numeric measurement value: {cell!r}",
                    })
                    continue

                below = False
                detection_limit = None
                value = cell
                if cell < 0:
                    below = True
                    detection_limit = abs(cell)
                    value = None

                rows_to_create.append(ReferenceSampleMeasurement(
                    import_ref=import_row,
                    reference_sample=sample,
                    element=element,
                    analytical_method=method,
                    value=value,
                    unit=element.default_unit,
                    below_detection_limit=below,
                    detection_limit=detection_limit,
                ))

    ReferenceSampleMeasurement.objects.bulk_create(
        rows_to_create,
        batch_size=2000,
        ignore_conflicts=True,
    )
    stats["measurements"] = len(rows_to_create)
