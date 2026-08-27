"""Seed reference-library lookup tables."""

from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from ...importers.seed_data import ELEMENTS
from ...models import DepositClassification, Element, Mineral


class Command(BaseCommand):
    help = "Seed Element / Mineral / DepositClassification lookup tables."

    def add_arguments(self, parser):
        parser.add_argument(
            "--metadata",
            type=Path,
            default=None,
            help=(
                "Optional path to an OSNACA-Metadata*.xlsx file. "
                "If omitted, only the Element table is seeded."
            ),
        )

    def handle(self, *args, **options):
        elements_created = self._seed_elements()
        self.stdout.write(self.style.SUCCESS(
            f"Element: {elements_created} rows upserted."
        ))

        metadata_path: Path | None = options["metadata"]
        if metadata_path is None:
            self.stdout.write(
                "No --metadata given; skipping Mineral / DepositClassification."
            )
            return

        if not metadata_path.is_file():
            raise CommandError(f"Metadata file not found: {metadata_path}")

        wb = openpyxl.load_workbook(metadata_path, read_only=True, data_only=True)
        try:
            minerals = self._seed_minerals(wb)
            self.stdout.write(self.style.SUCCESS(
                f"Mineral: {minerals} rows upserted."
            ))

            classifications = self._seed_classifications(wb)
            self.stdout.write(self.style.SUCCESS(
                f"DepositClassification: {classifications} rows upserted."
            ))
        finally:
            wb.close()

    def _seed_elements(self) -> int:
        count = 0
        for symbol, name, atomic_number, default_unit in ELEMENTS:
            Element.objects.update_or_create(
                symbol=symbol,
                defaults={
                    "name": name,
                    "atomic_number": atomic_number,
                    "default_unit": default_unit,
                },
            )
            count += 1
        return count

    def _seed_minerals(self, wb) -> int:
        ws = wb["Mineral Codes"]
        count = 0
        for name, code, *_ in ws.iter_rows(min_row=2, values_only=True):
            if not code or not name:
                continue
            Mineral.objects.update_or_create(code=code, defaults={"name": name})
            count += 1
        return count

    def _seed_classifications(self, wb) -> int:
        ws = wb["Ore Deposit Classification"]
        count = 0
        current_class = None
        for cls, sub, notes in ws.iter_rows(min_row=2, values_only=True):
            if cls:
                current_class = cls
            if current_class is None:
                continue
            DepositClassification.objects.update_or_create(
                deposit_class=current_class,
                sub_class=sub or "",
                defaults={"notes": notes or ""},
            )
            count += 1
        return count
